# Скрипт для получения данных из Google Sheets
# Необходимо в settings.ini указать
# * DB.TYPE - тип базы данных (куда выгружать данные)
# * DB.HOST - адрес (хост) базы данных
# * DB.PORT - порт хоста базы данных (если отличается от стандартного)
# * DB.USER - пользователь базы данных
# * DB.PASSWORD - пароль к базе данных
# * DB.DB - имя базы данных
# * GOOGLE_SHEETS.KEYS - ключи документов Google Sheets (через запятую)
# * GOOGLE_SHEETS.TABLES - таблицы для загрузки данных документов (через запятую в том же порядке)

# импорт общих библиотек
from datetime import datetime as dt
from datetime import date, timedelta
import pandas as pd
import numpy as np
import requests
from openpyxl import load_workbook
from io import StringIO,BytesIO
from sqlalchemy import create_engine, text
from requests.packages.urllib3.exceptions import InsecureRequestWarning
# Скрытие предупреждения Unverified HTTPS request
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
# Скрытие предупреждение про fillna
try:
    pd.set_option("future.no_silent_downcasting", True)
except Exception as E:
    pass

# импорт настроек
import configparser
config = configparser.ConfigParser()
config.read("../settings.ini")

# подключение к БД
if config["DB"]["PORT"] != '':
    DB_PORT = ':' + config["DB"]["PORT"]
else:
    DB_PORT = ''
if config["DB"]["TYPE"] == "MYSQL":
    engine = create_engine('mysql+mysqldb://' + config["DB"]["USER"] + ':' + config["DB"]["PASSWORD"] + '@' + config["DB"]["HOST"] + DB_PORT + '/' + config["DB"]["DB"] + '?charset=utf8')
elif config["DB"]["TYPE"] == "POSTGRESQL":
    engine = create_engine('postgresql+psycopg2://' + config["DB"]["USER"] + ':' + config["DB"]["PASSWORD"] + '@' + config["DB"]["HOST"] + DB_PORT + '/' + config["DB"]["DB"] + '?client_encoding=utf8')
elif config["DB"]["TYPE"] == "MARIADB":
    engine = create_engine('mariadb+mysqldb://' + config["DB"]["USER"] + ':' + config["DB"]["PASSWORD"] + '@' + config["DB"]["HOST"] + DB_PORT + '/' + config["DB"]["DB"] + '?charset=utf8')
elif config["DB"]["TYPE"] == "ORACLE":
    engine = create_engine('oracle+pyodbc://' + config["DB"]["USER"] + ':' + config["DB"]["PASSWORD"] + '@' + config["DB"]["HOST"] + DB_PORT + '/' + config["DB"]["DB"])
elif config["DB"]["TYPE"] == "SQLITE":
    engine = create_engine('sqlite:///' + config["DB"]["DB"])

# создание подключения к БД
if config["DB"]["TYPE"] in ["MYSQL", "POSTGRESQL", "MARIADB", "ORACLE", "SQLITE"]:
    connection = engine.connect()
    if config["DB"]["TYPE"] in ["MYSQL", "MARIADB"]:
        connection.execute(text('SET NAMES utf8mb4'))
        connection.execute(text('SET CHARACTER SET utf8mb4'))
        connection.execute(text('SET character_set_connection=utf8mb4'))

tables = config['GOOGLE_SHEETS']['TABLES'].split(',')
formats = config['GOOGLE_SHEETS']['FORMATS'].split(',')
for gs_i, gs_key in enumerate(config['GOOGLE_SHEETS']['KEYS'].split(',')):
    gs_key = gs_key.strip()
    gs_table = tables[gs_i].strip()
    if len (formats) >= gs_i-1:
        gs_format = formats[gs_i].strip()
    else:
        gs_format = 'CSV'
    if gs_format == 'CSV':
# формируем датафрейм из CSV ответа Google Sheets
        data = pd.read_csv(StringIO(requests.get('https://docs.google.com/spreadsheet/ccc?key=' + gs_key + '&output=csv').content.decode("utf-8")))
# базовый процесс очистки: приведение к нужным типам
        for col in data.columns:
# приведение строк
            data[col] = data[col].fillna('')
        import_data = {gs_table: data}
    elif gs_format == 'XLSX':
# получаем исходный документ и его листы
        excel_file = BytesIO(requests.get('https://docs.google.com/spreadsheets/d/e/' + gs_key + '/pub?output=xlsx').content)
        sheets = load_workbook(excel_file, read_only=True).sheetnames
        import_data = {}
# последовательно складываем все листы в данные и формируем список для загрузки
        for sheet in sheets:
            import_data[gs_table + '_' + sheet.replace(' ', '_')] = pd.read_excel(excel_file, engine='openpyxl', sheet_name=sheet)
    for table in import_data.keys():
        data = import_data[table]
        for col in data.columns:
            data[col] = data[col].fillna('').astype(str)
# поддержка TCP HTTP для Clickhouse
        if config["DB"]["PORT"] != '8443':
            CLICKHOUSE_PROTO = 'http://'
            CLICKHOUSE_PORT = config["DB"]["PORT"]
        else:
            CLICKHOUSE_PROTO = 'https://'
            CLICKHOUSE_PORT = config["DB"]["PORT"]
# создаем таблицу в первый раз
        if config["DB"]["TYPE"] == "CLICKHOUSE":
            requests.post(CLICKHOUSE_PROTO + config["DB"]["USER"] + ':' + config["DB"]["PASSWORD"] + '@' + config["DB"]["HOST"] + ':' + CLICKHOUSE_PORT + '/', verify=False,
                params={"database": config["DB"]["DB"], "query": (pd.io.sql.get_schema(data, table) + "  ENGINE=MergeTree ORDER BY (`" + list(data.columns)[0] + "`)").replace("CREATE TABLE ", "CREATE OR REPLACE TABLE " + config["DB"]["DB"] + ".").replace("INTEGER", "Int64")})
        if config["DB"]["TYPE"] in ["MYSQL", "POSTGRESQL", "MARIADB", "ORACLE", "SQLITE"]:
# обработка ошибок при добавлении данных
            try:
                data.to_sql(name=table, con=engine, if_exists='replace', chunksize=100)
            except Exception as E:
                print (E)
                connection.rollback()
        elif config["DB"]["TYPE"] == "CLICKHOUSE":
            csv_file = data.to_csv(index=False, header=False).encode('utf-8')
            requests.post(CLICKHOUSE_PROTO + config["DB"]["USER"] + ':' + config["DB"]["PASSWORD"] + '@' + config["DB"]["HOST"] + ':' + CLICKHOUSE_PORT + '/',
                params={"database": config["DB"]["DB"], "query": 'INSERT INTO ' + config["DB"]["DB"] + '."' + table + '" FORMAT CSV\n"' + '","'.join(data.columns) + '"\n"' + '","'.join(['String']*len(data.columns))},
                headers={'Content-Type':'application/octet-stream'}, data=csv_file, stream=True, verify=False)
        print (table + ":", len(data))

# закрытие подключения к БД
if config["DB"]["TYPE"] in ["MYSQL", "POSTGRESQL", "MARIADB", "ORACLE", "SQLITE"]:
    connection.commit()
    connection.close()